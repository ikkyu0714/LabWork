from googletrans import Translator
import openpyxl
from langrid.clients import TranslationClient
from langrid.settings import _config

gnmt = TranslationClient('http://langrid.org/service_manager/wsdl/kyoto1.langrid:GoogleTranslateNMT', 'sil.ritsumei', 'Shakaichinou')

EXCEL_PATH = '../../研究/WordNet比較データ/親子_所属単語比較_1216_2.xlsx'
book = openpyxl.load_workbook(EXCEL_PATH)
sheet = book['Sheet']

target_list = []

row = 1
for line in sheet.iter_rows(min_row=2):
    value = []
    print(row)
    row += 1
    for item in line:
        value.append(item.value)
    if value[1] == gnmt.translate("ja", "en", value[0].replace('_', ' ')):
        if value[0] == gnmt.translate("en", "ja", value[1].replace('_', ' ')):
            print('日本語:{}, 英語:{}, 翻訳:{}'.format(value[0], value[1].replace('_', ' '), gnmt.translate("ja", "en", value[0].replace('_', ' '))))
            print(value)
            target_list.append(value)
# 保存
book.save(EXCEL_PATH)
# 終了
book.close()

book_write = openpyxl.Workbook()
sheet1 = book_write.active

for data in target_list:
    max_row = sheet1.max_row
    sheet1.cell(row = max_row + 1, column = 1).value = data[0]
    sheet1.cell(row = max_row + 1, column = 2).value = data[1]
    sheet1.cell(row = max_row + 1, column = 3).value = data[2]
    sheet1.cell(row = max_row + 1, column = 4).value = data[3]
book_write.save('../../研究/WordNet比較データ/日英両方向_翻訳単語合致_1222.xlsx')
