import openpyxl
import os

class Questionaire_Transform():
    def __init__(self):
        self.list = []
        self.row_count = 1

    def add(self, data):
        self.list.append(data)

    def select_add(self, data, number):
        values = []
        if number % 3 == 1:
            values = [data[0], data[3], data[4], data[6], data[7], data[9], data[10]]
        elif number % 3 == 2:
            values = [data[2], data[5], data[3], data[8], data[6], data[9], data[10]]
        elif number % 3 == 0:
            values = [data[1], data[4], data[5], data[7], data[8], data[9], data[10]]
        else:
            print('else発動')
        self.list.append(values)

    def row_add(self):
        self.row_count += 4

    def row_reset(self):
        self.row_count = 1

    def print_list(self):
        for data in self.list:
            for item in data:
                print(item)

    def write_excel(self):
        book = openpyxl.Workbook()
        sheet = book.active
        row = 1
        for data in self.list:
            print(row)
            ques_row = 1
            for item in data:
                sheet.cell(row = row + ques_row, column = 2).value = item
                ques_row += 1
            row += ques_row - 1
        book.save('../../研究/アンケートデータ/まとめデータ/まとめ結果english.xlsx')

qgv = Questionaire_Transform()
books = '../../研究/アンケートデータ/日本データ抜き出し/english1結果.xlsx'
book = openpyxl.load_workbook(books)
sheet = book['Sheet']
number = 1
for line in sheet.iter_rows(min_row=2, min_col=2):
    values = []
    for item in line:
        values.append(item.value)
    qgv.select_add(values, number)
    number += 1
#qgv.print_list()
qgv.write_excel()

