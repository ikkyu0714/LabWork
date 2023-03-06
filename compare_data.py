"""
2021/12/08 研究用
日英インドネシア版と日英版のアンケートで変わったデータを調べる
差分を調べて、予算を見積もり
'synset_indonesiaのコピー2.xlsx'
'synsetアンケートのコピー.xlsx'
"""

import openpyxl


# 変数
max_height = [] # 各行の画像の高さの最大値を保持

class CompareData():
    def __init__(self):
        self.readfile1 = 'synset_indonesiaのコピー2.xlsx' # 読み込むsynsetのリストを記録したファイル
        self.readfile2 = 'synsetアンケートのコピー.xlsx' # 読み込むsynsetのリストを記録したファイル
        self.jpn_keyword_dict1 = {} # キーワードのリストを取得する辞書（日本語）
        self.eng_keyword_dict1 = {} # キーワードのリストを取得する辞書（英語）
        self.jpn_explanation_dict1 = {}
        self.eng_explanation_dict1 = {}
        self.jpn_keyword_dict2 = {} # キーワードのリストを取得する辞書（日本語）
        self.eng_keyword_dict2 = {} # キーワードのリストを取得する辞書（英語）
        self.jpn_explanation_dict2 = {}
        self.eng_explanation_dict2 = {}
        self.non_dict = {}

    # Synsetを取得する
    def get_synset(self, file_number):
        if file_number == 1:
            key = 1
            book = openpyxl.load_workbook(self.readfile1)
            sheet = book.worksheets[0]

            for row in sheet.iter_rows(min_row=2, min_col=2):
                value = []
                for cell in row:
                    values = cell.value
                    #print(values)
                    values = values.replace(',', '')
                    value.append(values.replace('_', ' '))

                self.jpn_keyword_dict1[key] = value[0]
                self.eng_keyword_dict1[key] = value[1]
                self.jpn_explanation_dict1[key] = value[3]
                self.eng_explanation_dict1[key] = value[4]
                key += 1
                
        if file_number == 2:
            key = 1
            book = openpyxl.load_workbook(self.readfile2)
            sheet = book.worksheets[0]

            for row in sheet.iter_rows(min_row=2, min_col=2):
                value = []
                for cell in row:
                    values = cell.value
                    #print(values)
                    values = values.replace(',', '')
                    value.append(values.replace('_', ' '))

                self.jpn_keyword_dict2[key] = value[0]
                self.eng_keyword_dict2[key] = value[1]
                self.jpn_explanation_dict2[key] = value[3]
                self.eng_explanation_dict2[key] = value[4]
                key += 1
                
    def compare_data(self):
        for i in range(1, 1001):
            if self.jpn_keyword_dict2[i] in self.jpn_keyword_dict1.values():
                if self.eng_keyword_dict2[i] in self.eng_keyword_dict1.values():
                    pass
                else:
                    self.non_dict[i] = self.jpn_keyword_dict2[i]
            else:
                self.non_dict[i] = self.jpn_keyword_dict2[i]

    def print_result(self):
        for key in self.non_dict.keys():
            print('データ:{}の{}'.format(key, self.non_dict[key]))
        print('データの総数:{}'.format(len(self.non_dict)))

cd = CompareData()
cd.get_synset(1)
cd.get_synset(2)
cd.compare_data()
cd.print_result()
