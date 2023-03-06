import openpyxl

def data_get(books): 
    book = openpyxl.load_workbook(books)
    sheet = book['Sheet']

    values = []

    for lines in sheet.iter_rows(min_row=3, min_col=1):
        item = [line.value for line in lines]
        values.append(item)

    return values

def write_result(results):
    book = openpyxl.Workbook()
    sheet = book.active
    id_count = 1
    sheet.cell(row = 1, column = 1).value = 'ID'
    sheet.cell(row = 1, column = 2).value = '文化差ラベル'
    for result in results:
        max_row = sheet.max_row # シートの最後の行を取得
        sheet.cell(row = max_row + 1, column = 1).value = id_count
        sheet.cell(row = max_row + 1, column = 2).value = result
        id_count += 1

    book.save('../../研究/アンケートデータ/文化差データ/文化差データ_多数決1なし.xlsx')

def judge(jpn_data, eng_data):
    judge_result = []
    for jpn_value, eng_value in zip(jpn_data, eng_data):
        # 日本人が日本語画像を想起できない、もしくはアメリカ人が英語画像を想起できない場合を外す
        if jpn_value[0] == '想起できない' or eng_value[1] == '想起できない':
            judge_result.append('対象外')

        # 日本人もアメリカ人もどちらも想起できる場合、文化差なし
        elif '想起できる' in jpn_value[1] and '想起できる' in eng_value[0]:
            judge_result.append('文化差なし')

        # お互いが相手の画像を想起できない場合、文化差ありの排他関係
        elif jpn_value[1] == '想起できない' and eng_value[0] == '想起できない':
            judge_result.append('文化差あり(排他関係)')

        # どちらか一方が想起できない場合、文化差ありの包含関係
        elif jpn_value[1] == '想起できない':
            judge_result.append('文化差あり(包含関係)')
        elif eng_value[0] == '想起できない':
            judge_result.append('文化差あり(包含関係)')

        else:
            judge_result.append('対象外')

    return judge_result

book_english = '../../研究/アンケートデータ/想起データ/english想起データ多数決_1なし.xlsx'
book_japanese = '../../研究/アンケートデータ/想起データ/japanese想起データ多数決_1なし.xlsx'
english_data = data_get(book_english)
japanese_data = data_get(book_japanese)

write_result(judge(japanese_data, english_data))