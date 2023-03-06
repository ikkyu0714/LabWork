import openpyxl

class CultureJudge():
    def __init__(self):
        self.judge_list = []
        self.japanese_list = []
        self.indonesian_list = []
        self.choice_list = []
        #self.searchword_list = ["object", "matter", "process", "causal_agent", "thing", "psychological_feature", "attribute", "group", "communication", "measure", "relation"]
        self.zerocount = 0
        self.count = 0

    # アメリカ人が英語の画像を想起できない, 日本人が日本語の画像を想起できない場合, データから外す
    def choicedata(self, data):
        if 0 in data: # データ欠損
            self.zerocount += 1
        elif data[3] == 1: # 日本人が想起できない
            self.count += 1
        elif data[8] == 1: # アメリカ人が想起できない
            self.count += 1
        else: # 使えるデータをリストに加える
            self.choice_list.append(data)
        """elif data[3] == 1 and data[5] == 1: # 日本人が想起できない
            self.count += 1
        elif data[8] == 1 and data[10] == 1: # アメリカ人が想起できない
            self.count += 1"""

    # 数値の値から想起できると想起できないを判定する
    def able_imagine(self):
        for line in self.choice_list:
            japanese = []
            value = 0
            japanese.append(line[0])
            # lineの中身(ID, 画像同士(日本), 画像同士(アメリカ), 説明と画像jpn(日本), 説明と画像jpn(アメリカ), キーワードと画像jpn(日本), キーワードと画像jpn(アメリカ), 説明と画像eng(日本), 説明と画像eng(アメリカ), キーワードと画像eng(日本), キーワードと画像eng(アメリカ), 説明とキーワード(日本), 説明とキーワード(アメリカ))
            if line[3] != 1 and line[5] != 1:
                value = (line[3] + line[5]) // 2
                japanese.append('想起できる({})'.format(value))
            elif line[3] == 1 or line[5] == 1:
                japanese.append('想起できない')
            else:
                japanese.append('想起できない')
            value = 0
            if line[7] != 1 and line[9] != 1:
                value = (line[7] + line[9]) // 2
                japanese.append('想起できる({})'.format(value))
            elif line[7] == 1 or line[9] == 1:
                japanese.append('想起できない')
            else:
                japanese.append('想起できない')
            self.japanese_list.append(japanese)
            value = 0
            indonesian = []
            indonesian.append(line[0])
            if line[4] != 1 and line[6] != 1:
                value = (line[4] + line[6]) // 2
                indonesian.append('想起できる({})'.format(value))
            elif line[4] == 1 or line[6] == 1:
                indonesian.append('想起できない')
            else:
                indonesian.append('想起できない')
            value = 0
            if line[8] != 1 and line[10] != 1:
                value = (line[8] + line[10]) // 2
                indonesian.append('想起できる({})'.format(value))
            elif line[8] == 1 or line[10] == 1:
                indonesian.append('想起できない')
            else:
                indonesian.append('想起できない')
            """if line[3] == 1 and line[5] == 1:
                japanese.append('想起できない')
            elif line[3] != 1 or line[5] != 1:
                value = (line[3] + line[5]) // 2
                japanese.append('想起できる({})'.format(value))
            else:
                japanese.append('想起できない')
            value = 0
            if line[7] == 1 and line[9] == 1:
                japanese.append('想起できない')
            elif line[7] != 1 or line[9] != 1:
                value = (line[7] + line[9]) // 2
                japanese.append('想起できる({})'.format(value))
            else:
                japanese.append('想起できない')
            self.japanese_list.append(japanese)
            value = 0
            indonesian = []
            indonesian.append(line[0])
            if line[4] == 1 and line[6] == 1:
                indonesian.append('想起できない')
            elif line[4] != 1 or line[6] != 1:
                value = (line[4] + line[6]) // 2
                indonesian.append('想起できる({})'.format(value))
            else:
                indonesian.append('想起できない')
            value = 0
            if line[8] == 1 and line[10] == 1:
                indonesian.append('想起できない')
            elif line[8] != 1 or line[10] != 1:
                value = (line[8] + line[10]) // 2
                indonesian.append('想起できる({})'.format(value))
            else:
                indonesian.append('想起できない')"""
            self.indonesian_list.append(indonesian)

    # 日本人とアメリカ人の想起レベルのマッチアップに応じて文化差を判定する
    def judgeculture(self):
        for jpn, ind in zip(self.japanese_list, self.indonesian_list):
            list = []
            if jpn[0] == ind[0]:
                list.append(jpn[0])
                if '想起できる' in jpn[1] and '想起できる' in ind[1]:
                    if '想起できる' in jpn[2] and '想起できる' in ind[2]:
                        list.append('文化差なし')
                    elif '想起できない' in jpn[2]:
                        print('日本:{}, アメリカ:{}'.format(jpn, ind))
                        list.append('文化差あり(包含関係)')
                    elif '想起できない' in ind[2]:
                        print('日本:{}, アメリカ:{}'.format(jpn, ind))
                        list.append('文化差あり(包含関係)')
                    else:
                        print('おかしい')
                        exit()
                elif '想起できない' in jpn[1]:
                    if '想起できる' in jpn[2] and '想起できる' in ind[2]:
                        print('日本:{}, アメリカ:{}'.format(jpn, ind))
                        list.append('文化差あり(包含関係)')
                    elif '想起できない' in jpn[2]:
                        print('日本:{}, アメリカ:{}'.format(jpn, ind))
                        list.append('文化差あり(包含関係)')
                    elif '想起できない' in ind[2]:
                        list.append('文化差あり(排他関係)')
                    else:
                        print('おかしい')
                        exit()
                elif '想起できない' in ind[1]:
                    if '想起できる' in jpn[2] and '想起できる' in ind[2]:
                        print('日本:{}, アメリカ:{}'.format(jpn, ind))
                        list.append('文化差あり(包含関係)')
                    elif '想起できない' in jpn[2]:
                        list.append('文化差あり(排他関係)')
                    elif '想起できない' in ind[2]:
                        print('日本:{}, アメリカ:{}'.format(jpn, ind))
                        list.append('文化差あり(包含関係)')
                    else:
                        print('おかしい')
                        exit()
                else:
                    print('おかしい')
                    exit()
            else:
                print('おかしいよ')
                exit()
            self.judge_list.append(list)

    def write_excel(self):
        book = openpyxl.Workbook()
        sheet = book.active
        row = 1
        number = 0
        while row <= 1000:
            data = self.judge_list[number]
            if data[0] == row:
                sheet.cell(row = row, column = 1).value = data[1]
                number += 1
                row += 1
            else:
                row += 1
        book.save('人手判定(日英)_1013.xlsx')

    def print_choicedata(self):
        for line in self.choice_list:
            print(line)
        print(len(self.choice_list))

    def print_count(self):
        print(self.zerocount)
        print(self.count)

    def printjudge(self):
        includecount = 0
        exclusivecount = 0
        nocount = 0
        for line in self.judge_list:
            if '包含' in line[1]:
                includecount += 1
            elif '排他' in line[1]:
                exclusivecount += 1
            elif line[1] == '文化差なし':
                nocount += 1
            print(line)
        print('排他関係:{}, 包含関係:{}, 文化差なし:{}, 文化差あり:{}'.format(exclusivecount, includecount, nocount, exclusivecount+includecount))

cj = CultureJudge()
books = '日本アメリカやり直し.xlsx'
book = openpyxl.load_workbook(books)
sheet = book['Sheet']
max_row = sheet.max_row + 1
number = 1
row_start = 2
row_stop = 8
while row_stop <= max_row:
    question = []
    question.append(number)
    for line in sheet.iter_rows(min_row=row_start, max_row=row_stop):
        values = []
        for item in line:
            values.append(item.value)
        question.extend(values)
    cj.choicedata(question)
    row_start += 7
    row_stop += 7
    number += 1
cj.print_choicedata()
cj.print_count()
cj.able_imagine()
cj.judgeculture()
cj.printjudge()
#cj.write_excel()