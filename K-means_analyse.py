# -*- coding: utf-8 -*-
"""
# coding:utf-8
2021/11/18 研究用
従来手法とK-meansを使った手法, VGG16の予測を使った手法を比較分析
"""
import openpyxl
import random
from def_sim_data import Excel_Data
from search_excel import Search_Excel
from openpyxl.styles import PatternFill

class SuperExcelData(Excel_Data):
    def __init__(self, id, japanese, english, answer, similarity, cla2, cla3, cla4, hypernym):
        super().__init__(id, japanese, english, answer, similarity, hypernym)
        self.cla2 = cla2
        self.cla3 = cla3
        self.cla4 = cla4

class CompareVggKmeans(Search_Excel):
    def __init__(self):
        super().__init__()
        self.testdata = []
        self.thre_list = [0.0, 0.05, 0.1, 0.15, 0.2, 0.25, 0.3, 0.35, 0.4, 0.45, 0.5, 0.55, 0.6, 0.65, 0.7, 0.75, 0.8, 0.85, 0.9, 0.95, 1.0]
        self.oricount = []
        self.cla2count = []
        self.cla3count = []
        self.cla4count = []
        self.datacount = []
        self.both = []
        self.culture = []
        self.ori_thre = 0.55
        self.cla2_thre = 0.55
        self.cla3_thre = 0.5
        self.cla4_thre = 0.5
        self.bothpositive_id_list = []
        self.bothnegative_id_list = []
        self.abpositive_id_list = []
        self.abnegative_id_list = []

    def compare(self):
        count = 0
        datacount = 0
        for line in self.list:
            #print(line.answer)
            if line.answer == '文化差なし':
                datacount += 1
                if line.similarity >= self.ori_thre and line.cla2 < self.cla2_thre:
                    count += 1
                    self.bothnegative_id_list.append(line.id)
                elif line.similarity < self.ori_thre and line.cla2 >= self.cla2_thre:
                    count += 1
                    self.bothpositive_id_list.append(line.id)
            
            elif '文化差あり' in line.answer:
                datacount += 1
                if line.similarity < self.ori_thre and line.cla2 >= self.cla2_thre:
                    count += 1
                    self.abnegative_id_list.append(line.id)
                elif line.similarity >= self.ori_thre and line.cla2 < self.cla2_thre:
                    count += 1
                    self.abpositive_id_list.append(line.id)

    def writeresult(self):
        file = 'compareclaster_result.xlsx'
        book = openpyxl.load_workbook(file)
        sheet = book.worksheets[0]
        
        # 最後の一つ下の行に (3列目:日本語のキーワード, 4列目:英語のキーワード, 5列目:cos類似度) を書き込む
        for thre, simi, vgg, kmeans, data in zip(self.thre_list, self.oricount, self.vggcount, self.kmeanscount, self.datacount):
            max_row = sheet.max_row # シートの最後の行を取得
            #print('閾値:{}, 従来:{}%, VGG:{}%, Kmeans:{}%, data:{}'.format(thre, round(simi/data*100, 2), round(vgg/data*100,2), round(kmeans/data*100,2), data))
            sheet.cell(row = max_row + 1, column = 1).value = thre
            sheet.cell(row = max_row + 1, column=2).value = round(simi/data*100, 2)
            sheet.cell(row = max_row + 1, column=3).value = round(vgg/data*100,2)
            sheet.cell(row = max_row + 1, column=4).value = round(kmeans/data*100,2)
        # 保存
        book.save(file)
        # 終了
        book.close()
        
    def add_simi(self):
        simi = 0
        cla2 = 0
        for line in self.list:
            simi += line.similarity
            cla2 += line.cla2
        print('類似度:{}, K-means:{}'.format(simi/500, cla2/500))
    
    def coloring(self, id_list, color):
        for list in id_list:
            for row in sheet.iter_rows():
                for cell in row:
                    if row[0].value == list:
                        cell.fill = PatternFill(fgColor=color,bgColor=color, fill_type = "solid")

    def resultprint(self):
        print('両方')
        print('===============positive=================')
        for id in self.bothpositive_id_list:
            print('ID:{}'.format(id))
        print('===============positive=================')
        print('===============negative=================')
        for id in self.bothnegative_id_list:
            print('ID:{}'.format(id))
        print('===============negative=================')
        print('AB')
        print('===============positive=================')
        for id in self.abpositive_id_list:
            print('ID:{}'.format(id))
        print('===============positive=================')
        print('===============negative=================')
        for id in self.abnegative_id_list:
            print('ID:{}'.format(id))
        print('===============negative=================')
        print('ポジティブの総数{}, ネガティブの総数{}'.format(len(self.bothpositive_id_list), len(self.bothnegative_id_list)))
        print('ポジティブの総数{}, ネガティブの総数{}'.format(len(self.abpositive_id_list), len(self.abnegative_id_list)))

# エクセルファイルをロード
EXCEL_PATH = 'new_datasetやり直し0513.xlsx'
book = openpyxl.load_workbook(EXCEL_PATH)
sheet = book['Sheet1']

# インスタンスを生成
to = CompareVggKmeans()
"""
'339966':緑
'FF0000':赤
'3366FF':青
'FFFF00':黄色
"""
color = ['339966', 'FF0000', '3366FF', 'FFFF00']

for line in sheet.iter_rows(min_row=3):
    values = []
    for item in line:
        values.append(item.value)
    #print(values[0])
    to.add(SuperExcelData(values[0], values[1], values[2], values[3], values[4], values[5], values[6], values[7], values[8]))
#to.dataset()
to.compare()
to.coloring(to.bothpositive_id_list, color[0])
to.coloring(to.bothnegative_id_list, color[1])
to.coloring(to.abpositive_id_list, color[2])
to.coloring(to.abnegative_id_list, color[3])
to.resultprint()
to.add_simi()
#to.writeresult()

# 保存
book.save(EXCEL_PATH)
# 終了
book.close()
