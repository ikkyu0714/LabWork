import openpyxl
import matplotlib.pyplot as plt

EXCEL_PATH = '../../研究/WordNet比較データ/親子_所属単語比較_1216_2.xlsx'
book = openpyxl.load_workbook(EXCEL_PATH)
sheet = book['Sheet']

target_list = []

for line in sheet.iter_rows(min_row=2):
    value = []
    for item in line:
        value.append(item.value)
    target_list.append(value)
# 保存
book.save(EXCEL_PATH)
# 終了
book.close()
same_pare_list = []

same_count_list = []

for target in target_list:
    same_count_list.append(int(target[4]))

print('個数{}'.format(same_count_list))
plt.hist(same_count_list)
plt.show()
#print('全体の個数{}'.format(len(target_list)))

"""for i in range(len(target_list)):
    for j in range(i+1, len(target_list), 1):
        if target_list[i][0] == target_list[j][0] and target_list[i][1] == target_list[j][1]:
            same_pare_list.append([target_list[i], target_list[j]])
            #print('target:{}, 探す:{}'.format(target_list[i], target_list[j]))"""


"""book_write = openpyxl.Workbook()
sheet = book_write.active
print(len(same_pare_list))

for data in same_pare_list:
    max_row = sheet.max_row
    sheet.cell(row = max_row + 1, column = 1).value = data[0][0]
    sheet.cell(row = max_row + 1, column = 2).value = data[0][1]
    sheet.cell(row = max_row + 1, column = 3).value = data[0][2]
    sheet.cell(row = max_row + 1, column = 4).value = data[0][3]
    sheet.cell(row = max_row + 1, column = 5).value = data[0][4]
    sheet.cell(row = max_row + 1, column = 6).value = data[0][5]
    sheet.cell(row = max_row + 1, column = 7).value = data[1][2]
    sheet.cell(row = max_row + 1, column = 8).value = data[1][3]
    sheet.cell(row = max_row + 1, column = 9).value = data[1][4]
    sheet.cell(row = max_row + 1, column = 10).value = data[1][5]

book_write.save('../../研究/WordNet比較データ/親子兄弟_同一単語ペア_1216.xlsx')"""


