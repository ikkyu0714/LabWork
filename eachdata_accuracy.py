"""
2021/12/21 研究用
従来手法とK-means, 外れ値検出, K-means+外れ値検出(最適な閾値, 0.55, 0.6, 0.5, 0.5)
dataset_isolation.xlsx
従来手法とK-means, DBSCAN(1000), DBSCAN(1500)(最適な閾値, 0.55, 0.55, 0.55, 0.55)
dataset_dbscan.xlsx
4つの手法で各データ(文化差ありと文化差なし)の精度を調べる
"""

import openpyxl
import random
from def_sim_data import Excel_Data
from search_excel import Search_Excel

class SuperExcelData(Excel_Data):
    def __init__(self, id, japanese, english, answer, similarity, kmeans, isolation, kmeans_iso, hypernym):
        super().__init__(id, japanese, english, answer, similarity, hypernym)
        self.kmeans = kmeans
        self.isolation = isolation
        self.kmeans_iso = kmeans_iso

class EachDataAccuracy(Search_Excel):
    def __init__(self):
        super().__init__()
        self.testdata = []
        self.thre_list = [0.0, 0.05, 0.1, 0.15, 0.2, 0.25, 0.3, 0.35, 0.4, 0.45, 0.5, 0.55, 0.6, 0.65, 0.7, 0.75, 0.8, 0.85, 0.9, 0.95, 1.0]
        self.oricount = []
        self.kmeanscount = []
        self.isolationcount = []
        self.kmeans_isocount = []
        self.datacount = []
        self.both = []
        self.culture = []
    
    def dataset(self):
        # 文化差ありデータと文化差なしデータに分割する
        for line in self.list:
            #self.setdata.append(line)
            if line.answer == '両方':
                self.both.append(line)
            elif line.answer == 'AB':
                self.culture.append(line)
        
        # 上位250個ずつをテストデータにセット
        self.testdata = self.both[:250]
        self.testdata.extend(self.culture[:250])
        
    def data_excel_write(self):
        file = 'dataset.xlsx'
        book = openpyxl.load_workbook(file)
        sheet = book.worksheets[0]
        for line in self.testdata:
            max_row = sheet.max_row # シートの最後の行を取得
            sheet.cell(row = max_row + 1, column = 1).value = line.id
            sheet.cell(row = max_row + 1, column = 2).value = line.japanese
            sheet.cell(row = max_row + 1, column = 3).value = line.english
            sheet.cell(row = max_row + 1, column = 4).value = line.answer
            sheet.cell(row = max_row + 1, column = 5).value = line.similarity
            sheet.cell(row = max_row + 1, column = 6).value = line.vgg
            sheet.cell(row = max_row + 1, column = 7).value = line.kmeans
            sheet.cell(row = max_row + 1, column = 8).value = line.isolation
            sheet.cell(row = max_row + 1, column = 9).value = line.kmeansisolation
            sheet.cell(row = max_row + 1, column = 10).value = line.hypernym
            print('{},{},{},{},{},{}'.format(line.id, line.japanese, line.english, line.answer, line.similarity, line.kmeans, line.isolation, line.kmeansisolation))
        # 保存
        book.save(file)
        # 終了
        book.close()

    def compare(self):
        count = 0
        count_cul = 0
        countkmeans = 0
        kmeans_cul = 0
        countisolation = 0
        isolation_cul = 0
        countkmeans_iso = 0
        kmeans_iso_cul = 0
        datacount = 0
        for line in self.testdata:
            if line.answer == '両方':
                datacount += 1
                if line.similarity >= 0.55:
                    count += 1
                if line.kmeans >= 0.55:
                    countkmeans += 1
                if line.isolation >= 0.55:
                    countisolation += 1
                if line.kmeans_iso >= 0.55:
                    countkmeans_iso += 1
            elif line.answer == 'AB':
                datacount += 1
                if line.similarity < 0.55:
                    count_cul += 1
                if line.kmeans < 0.55:
                    kmeans_cul += 1
                if line.isolation < 0.55:
                    isolation_cul += 1
                if line.kmeans_iso < 0.55:
                    kmeans_iso_cul += 1
        self.oricount.append(count+count_cul)
        self.kmeanscount.append(countkmeans+kmeans_cul)
        self.isolationcount.append(countisolation+isolation_cul)
        self.kmeans_isocount.append(countkmeans_iso+kmeans_iso_cul)
        self.datacount.append(datacount)
        self.oricount.append(count)
        self.kmeanscount.append(countkmeans)
        self.isolationcount.append(countisolation)
        self.kmeans_isocount.append(countkmeans_iso)
        self.datacount.append(datacount)
        self.oricount.append(count_cul)
        self.kmeanscount.append(kmeans_cul)
        self.isolationcount.append(isolation_cul)
        self.kmeans_isocount.append(kmeans_iso_cul)
        self.datacount.append(datacount)

    def writeresult(self):
        file = 'comparemethod_result.xlsx'
        book = openpyxl.load_workbook(file)
        sheet = book.worksheets[0]
        titles = ['全体', '文化差なし', '文化差あり']

        # 最後の一つ下の行に (3列目:日本語のキーワード, 4列目:英語のキーワード, 5列目:cos類似度) を書き込む
        for title, thre, simi, kmeans, iso, kmeans_iso, data in zip(titles, self.thre_list, self.oricount, self.kmeanscount, self.isolationcount, self.kmeans_isocount, self.datacount):
            max_row = sheet.max_row # シートの最後の行を取得
            #print('閾値:{}, 従来:{}%, VGG:{}%, Kmeans:{}%, data:{}'.format(thre, round(simi/data*100, 2), round(vgg/data*100,2), round(kmeans/data*100,2), data))
            sheet.cell(row = max_row + 1, column = 1).value = thre
            sheet.cell(row = max_row + 1, column=2).value = round(simi/data*100, 2)
            sheet.cell(row = max_row + 1, column=3).value = round(kmeans/data*100,2)
            sheet.cell(row = max_row + 1, column=4).value = round(iso/data*100, 2)
            sheet.cell(row = max_row + 1, column=5).value = round(kmeans_iso/data*100,2)
        # 保存
        book.save(file)
        # 終了
        book.close()

    def resultprint(self):
        titles = ['全体', '文化差なし', '文化差あり']
        for title, simi, kmeans, iso, kmeans_iso, data in zip(titles, self.oricount, self.kmeanscount, self.isolationcount, self.kmeans_isocount, self.datacount):
            print('{}のAccuracy 従来:{}%, Kmeans:{}%, Isolation:{}%, Kmeans+Isolation:{}%, data:{}'.format(title, round(simi/data*100, 2), round(kmeans/data*100,2), round(iso/data*100,2), round(kmeans_iso/data*100,2), data))

# エクセルファイルをロード
EXCEL_PATH = 'dataset_isolation.xlsx'
book = openpyxl.load_workbook(EXCEL_PATH)
sheet = book['Sheet1']

# インスタンスを生成
to = EachDataAccuracy()
print('-------------------{}------------------'.format(EXCEL_PATH))
for line in sheet.iter_rows(min_row=2):
    values = []
    for item in line:
        values.append(item.value)
    to.add(SuperExcelData(values[0], values[1], values[2], values[3], values[4], values[5], values[6], values[7], values[8]))
to.dataset()
#to.data_excel_write()
to.compare()
to.resultprint()
#to.writeresult()

# 保存
book.save(EXCEL_PATH)
# 終了
book.close()
