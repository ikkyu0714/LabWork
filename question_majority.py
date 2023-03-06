"""
研究用 2022/10/17
アンケート結果の値を平均するプログラム
"""

import openpyxl
import os
import collections

class AverageQuestion():
    def __init__(self):
        self.average_list = []
        self.list = []

    def write_excel(self):
        book = openpyxl.Workbook()
        sheet = book.active
        row = 2
        sheet.cell(row = 1, column = 1).value = '日本人'
        for data in self.average_list:
            if data == 'スペース':
                row += 1
            else:
                sheet.cell(row = row, column = 1).value = data
                row += 1

        book.save('../../研究/アンケートデータ/平均データ/日本人_多数決平均.xlsx')

    # リストの中の最頻値を返す。もし最頻値が複数あるなら平均に回す
    def majority(self, list):

        # 0を取り除く
        list = [x for x in list if x != 0]

        # リストの中の頻出回数を辞書型で記憶
        count_num_dict = collections.Counter(list)

        # 頻出回数の最大をキーと値を多次元配列で格納
        max_list = [value for value in count_num_dict.items() if value[1] == max(count_num_dict.values())]

        # 最頻値が複数ある場合、値を平均して返す ※round関数は四捨五入ではなくただ丸めているだけ 1.5のように中間なら1と2の偶数の方になる
        if len(max_list) != 1:
            return round(sum(list)/len(list))
        # 最頻値が一つの場合、その最頻出の値を返す
        else:
            return max_list[0][0]

    def data_average(self, list):
        person_num = len(list) // 7 # アンケート回答者の人数
        array_list = [[] for i in range(person_num)]

        # 回答者ごとに回答を分別する
        for num in range(len(list)):
            person = int(num % person_num) # 何人目の回答者の回答かを判別する値
            array_list[person].append(list[num])

        # Yes以外の回答(No)があった時、その回答は信用できないため、Noの回答者の値を0にする
        for target in range(len(array_list)):
            value = []
            if array_list[target][6] == 'yes' or array_list[target][6] == 'Yes':
                value.extend(array_list[target])
            else:
                value.extend([0,0,0,0,0,0,'No'])
            array_list[target] = value

        # 質問ごとに回答者全員の値を取り出し、平均する
        for i in range(7): # 質問数が7個のため7回
            values = [value[i] for value in array_list]

            # 数字なら平均、文字ならスペースにする
            if type(values[0]) == int:
                ave = self.majority(values)
            else:
                ave = 'スペース'

            # 結果を保存
            self.average_list.append(ave)

    # 使うデータを選ぶ もし, Noが入っている場合、その概念におけるNoの回答者の回答は使用しない
    def choice_data(self, data):
        person_num = len(data) // 7
        array_list = [[] for i in range(person_num)]
        values = []
        for item in data:
            if type(item) == str:
                if ' ' in item:
                    item.replace(' ','')
                if '\n' in item:
                    item.replace('\n','')
            values.append(item)

        data = values
        self.data_average(data)
        #self.list.append(data)

aq = AverageQuestion()
books = '../../研究/アンケートデータ/まとめデータ/まとめ結果japanese一覧.xlsx'
book = openpyxl.load_workbook(books)
sheet = book['Sheet']
max_row = sheet.max_row # 最後の行を取得
number = 1 # 何番目かをカウント
row_start = 2 # 取得する最初の行
row_stop = 8 # 取得する最後の行
list = []
average_list = []

# 最初の行から最後の行を取得する １質問ずつ取得
while row_stop <= max_row:
    question = [] # 一つの概念の質問をまとめるリスト

    # startからstopまでの行をスライス
    for line in sheet.iter_rows(min_row=row_start, max_row=row_stop):
        values = [] # 1行の値をまとめるリスト

        # 行内の値を取り出してvaluesに入れる
        for item in line:
            values.append(item.value)
        question.extend(values)
    #print(row_start)
    aq.choice_data(question)
    row_start += 7
    row_stop += 7
    number += 1
aq.write_excel()

