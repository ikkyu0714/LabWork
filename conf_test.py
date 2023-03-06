"""
研究用
2021/06/03
全結合層ありとなしの比較
"""

import csv
import openpyxl
import pprint
from csv_reader import CSV_Reader
import random

class UnderSample_Compare(CSV_Reader):
    def __init__(self):
        super().__init__()
        self.conf0_list = [] # 類似度を入れるインスタンス変数
        self.conf1_list = [] # 人の判定結果を入れるインスタンス変数
        self.id_list = [] # IDを入れるインスタンス変数
        self.count_conf1 = []
        self.count_conf0 = []
        self.simi_result = []
        self.ampere_list = []
        self.ampere_float_list = []
        self.ampere_result = []
        self.write_data = [] # 書き込むデータ(ID, Answer, 類似度, Ampareさんの類似度)を入れるインスタンス変数
        self.answer_change_list = [] # 人の判定結果をもとに0 or 1で表記したものを入れるインスタンス変数
        self.culture_list = [] # 文化差ありデータ
        self.not_culture_list = [] # 文化差なしデータ
        self.test_data = []
        self.thre_list = [0.0, 0.05, 0.1, 0.15, 0.2, 0.25, 0.3, 0.35, 0.4, 0.45, 0.5, 0.55, 0.6, 0.65, 0.7, 0.75, 0.8, 0.85, 0.9, 0.95, 1.0]

    # データをCSVファイルから取得する
    def get_data(self):
        for data in self.fileread[1:]:
            self.id_list.append(data[0])
            self.conf1_list.append(data[1])
            self.conf0_list.append(data[2])

    # テスト(最適化)
    def test_optimize(self):
        for conf in self.conf1_list:
            for count in range(len(self.thre_list)-1):
                print(count)
                if self.thre_list[count] <= float(conf) and self.thre_list[count+1] > float(conf):
                    self.count_conf1[count] = 1 + self.count_conf1[count]

        for conf in self.conf0_list:
            for count2 in range(len(self.thre_list)-1):
                print(count2)
                if self.thre_list[count2] <= float(conf) and self.thre_list[count2+1] > float(conf):
                    self.count_conf0[count2] = 1 + self.count_conf0[count2]

        print(self.count_conf1)
        print(self.count_conf0)

    def write_excel(self):
        book = openpyxl.load_workbook('include_top_compare.xlsx')
        sheet = book['Sheet1']
        row = 1

        for data in self.test_data:
            sheet.cell(row = row, column = 1).value = data[0]
            sheet.cell(row = row, column = 2).value = data[1]
            sheet.cell(row = row, column = 3).value = data[2]
            sheet.cell(row = row, column = 4).value = data[3]
            row += 1

        book.save('include_top_compare.xlsx')
        
        book2 = openpyxl.load_workbook('compare_thre.xlsx')
        sheet2 = book2['Sheet1']
        row = 1

        for thre, ikkyu, ampere in zip(self.thre_list, self.simi_result, self.ampere_result):
            sheet2.cell(row = row, column = 1).value = thre
            sheet2.cell(row = row, column = 2).value = ikkyu/len(self.test_data)
            sheet2.cell(row = row+25, column = 1).value = thre
            sheet2.cell(row = row+25, column = 2).value = ampere/len(self.test_data)
            row += 1
        #print(data[1])

        book2.save('compare_thre.xlsx')

    def print_accuracy(self):
        message = 'Accuracyは{}, アンペアさんのAccuracyは{}です。'
        for result1, result2, threshold in zip(self.simi_result, self.ampere_result, self.thre_list):
            print('閾値:{}, Accuracy：{}, Ampere:{}'.format(threshold, result1/len(self.test_data), result2/len(self.test_data)))
        print(message.format(self.acc_ikkyu, self.acc_ampere))
        print(len(self.test_data))

filename = 'confidence_test.csv'
undersample_compare = UnderSample_Compare()
undersample_compare.read(filename)
undersample_compare.get_data()

#undersample_compare.data_set()
#undersample_compare.test()
undersample_compare.test_optimize()
#undersample_compare.print_accuracy()
#undersample_compare.write_excel()
