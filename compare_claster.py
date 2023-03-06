"""
2021/10/06 研究用
従来手法とK-meansを使った手法, VGG16の予測を使った手法を比較する
"""

import openpyxl
import random
from def_sim_data import Excel_Data
from search_excel import Search_Excel

class SuperExcelData(Excel_Data):
    def __init__(self, id, japanese, english, answer, similarity, claster2, claster3, claster4, hypernym):
        super().__init__(id, japanese, english, answer, similarity, hypernym)
        self.claster2 = claster2
        self.claster3 = claster3
        self.claster4 = claster4

class CompareVggKmeans(Search_Excel):
    def __init__(self):
        super().__init__()
        self.testdata = []
        self.thre_list = [0.0, 0.05, 0.1, 0.15, 0.2, 0.25, 0.3, 0.35, 0.4, 0.45, 0.5, 0.55, 0.6, 0.65, 0.7, 0.75, 0.8, 0.85, 0.9, 0.95, 1.0]
        self.oricount = []
        self.cla2count = []
        self.cla3count = []
        self.cla4count = []
        self.datacount = []
        self.both = []
        self.culture = []
    
    def dataset(self):
        # 文化差ありデータと文化差なしデータに分割する
        for line in self.list:
            #self.setdata.append(line)
            if line.answer == '両方':
                self.both.append(line)
            elif line.answer == 'AB':
                self.culture.append(line)
        
        # シャッフル
        random.shuffle(self.both)
        random.shuffle(self.culture)
        
        # 上位250個ずつをテストデータにセット
        self.testdata = self.both[:250]
        self.testdata.extend(self.culture[:250])
        
    def data_excel_write(self):
        file = 'dataset_dbscan.xlsx'
        book = openpyxl.load_workbook(file)
        sheet = book.worksheets[0]
        for line in self.testdata:
            max_row = sheet.max_row # シートの最後の行を取得
            sheet.cell(row = max_row + 1, column = 1).value = line.id
            sheet.cell(row = max_row + 1, column = 2).value = line.japanese
            sheet.cell(row = max_row + 1, column = 3).value = line.english
            sheet.cell(row = max_row + 1, column = 4).value = line.answer
            sheet.cell(row = max_row + 1, column = 5).value = line.similarity
            sheet.cell(row = max_row + 1, column = 6).value = line.claster2
            sheet.cell(row = max_row + 1, column = 7).value = line.claster3
            sheet.cell(row = max_row + 1, column = 8).value = line.claster4
            sheet.cell(row = max_row + 1, column = 9).value = line.hypernym
            #print('{},{},{},{},{},{},{},{}'.format(line.id, line.japanese, line.english, line.answer, line.similarity, line.claster2, line.claster3, line.claster4))
        # 保存
        book.save(file)
        # 終了
        book.close()

    def compare(self):
        for thre in self.thre_list:
            count = 0
            countcla2 = 0
            countcla3 = 0
            countcla4 = 0
            datacount = 0
            for line in self.testdata:
                if line.answer == '両方':
                    datacount += 1
                    if line.similarity >= thre:
                        count += 1
                    if line.claster2 >= thre:
                        countcla2 += 1
                    if line.claster3 >= thre:
                        countcla3 += 1
                    if line.claster4 >= thre:
                        countcla4 += 1
                elif line.answer == 'AB':
                    datacount += 1
                    if line.similarity < thre:
                        count += 1
                    if line.claster2 < thre:
                        countcla2 += 1
                    if line.claster3 < thre:
                        countcla3 += 1
                    if line.claster4 < thre:
                        countcla4 += 1
            self.oricount.append(count)
            self.cla2count.append(countcla2)
            self.cla3count.append(countcla3)
            self.cla4count.append(countcla4)
            self.datacount.append(datacount)

    def writeresult(self):
        file = 'dbscan_result.xlsx'
        book = openpyxl.load_workbook(file)
        sheet = book.worksheets[0]
        
        # 最後の一つ下の行に (3列目:日本語のキーワード, 4列目:英語のキーワード, 5列目:cos類似度) を書き込む
        for thre, simi, cla2, cla3, cla4, data in zip(self.thre_list, self.oricount, self.cla2count, self.cla3count, self.cla4count, self.datacount):
            max_row = sheet.max_row # シートの最後の行を取得
            #print('閾値:{}, 従来:{}%, VGG:{}%, Kmeans:{}%, data:{}'.format(thre, round(simi/data*100, 2), round(vgg/data*100,2), round(kmeans/data*100,2), data))
            sheet.cell(row = max_row + 1, column = 1).value = thre
            sheet.cell(row = max_row + 1, column=2).value = round(simi/data*100, 2)
            sheet.cell(row = max_row + 1, column=3).value = round(cla2/data*100,2)
            sheet.cell(row = max_row + 1, column=4).value = round(cla3/data*100,2)
            sheet.cell(row = max_row + 1, column=5).value = round(cla4/data*100,2)
        # 保存
        book.save(file)
        # 終了
        book.close()

    def resultprint(self):
        for thre, simi, cla2, cla3, cla4, data in zip(self.thre_list, self.oricount, self.cla2count, self.cla3count, self.cla4count, self.datacount):
            print('閾値:{}, 従来:{}%, ２クラスター:{}%, 3クラスター:{}%, 4クラスター:{}%, data:{}'.format(thre, round(simi/data*100, 2), round(cla2/data*100,2), round(cla3/data*100,2), round(cla4/data*100,2), data))

# エクセルファイルをロード
EXCEL_PATH = 'k-means-dbscan比較.xlsx'
book = openpyxl.load_workbook(EXCEL_PATH)
sheet = book['Sheet1']

# インスタンスを生成
to = CompareVggKmeans()

for line in sheet.iter_rows(min_row=3):
    values = []
    for item in line:
        values.append(item.value)
    to.add(SuperExcelData(values[0], values[1], values[2], values[3], values[4], values[5], values[6], values[7], values[8]))
to.dataset()
to.data_excel_write()
to.compare()
to.resultprint()
to.writeresult()

# 保存
book.save(EXCEL_PATH)
# 終了
book.close()
