"""
2022/10/13 研究用
Accuracyを計測する
"""

import openpyxl # エクセルファイルを使うモジュール
import random # ランダム関数のモジュール
from modules.def_sim_data import Excel_Data
from modules.search_excel import Search_Excel
import itertools

class SuperExcelData(Excel_Data):
    def __init__(self, id, japanese, english, answer, similarity, hypernym):
        super().__init__(id, japanese, english, answer, similarity, hypernym)

class AccuracyCalculate(Search_Excel):
    def __init__(self):
        super().__init__()
        self.testdata = [] # テストデータの入れ物
        self.validationdata = [] # バリデーションデータの入れ物

        # 5交差検証用のリスト
        self.data1 = []
        self.data2 = []
        self.data3 = []
        self.data4 = []
        self.data5 = []

        # 閾値のリスト
        self.thre_list = [0.0, 0.05, 0.1, 0.15, 0.2, 0.25, 0.3, 0.35, 0.4, 0.45, 0.5, 0.55, 0.6, 0.65, 0.7, 0.75, 0.8, 0.85, 0.9, 0.95, 1.0]

        self.oricount = [] # 正解数をカウントする
        self.datacount = [] # データ数をカウントする
        self.both = [] # 文化差なしデータを入れる
        self.culture = [] # 文化差ありデータを入れる
        self.validation_result = [] # バリデーションの結果を入れる
        self.test_result = [] # テストの結果を入れる

    # データを分割して入れる関数
    def data_split(self, start, stop):
        # 各ラベルのデータリストのうちstart番目からstop番目に入っているデータを入れる
        list = [] # データを入れるリスト
        list.extend(self.both[start:stop:1]) # 文化差なしデータを入れる
        list.extend(self.culture[start:stop:1]) # 文化差ありデータを入れる

        return list

    # 文化差ありデータと文化差なしデータに分割する
    def dataset(self):
        for line in self.list:
            # 回答者の回答が文化差ありかなしか空欄かを判断
            if line.answer == None:
                pass
            elif line.answer == '文化差なし': # 文化差なしデータ
                self.both.append(line)
            elif '文化差あり' in line.answer: # 文化差ありデータ
                self.culture.append(line)

        # IDの順番通りになっているためシャッフル
        random.shuffle(self.both)
        random.shuffle(self.culture)

        #############################
        # 使うデータの数
        len_data = 115
        #############################

        # 使用するデータを格納
        self.test_data = self.data_split(0, len_data)
        print(len(self.test_data))

        # データを５等分するために、データの数の５つの区切りをリストに格納 データ数100の場合, num_list = [20,40,60,80,100]
        num_list = []
        for num in range(1, 6, 1):
            num_list.append((len_data//5)*num)

        # 各データ群に'文化差なし'データと'文化差あり'データを同数ずつ入れる
        self.data1 = self.data_split(0, num_list[0])
        self.data2 = self.data_split(num_list[0], num_list[1])
        self.data3 = self.data_split(num_list[1], num_list[2])
        self.data4 = self.data_split(num_list[2], num_list[3])
        self.data5 = self.data_split(num_list[3], num_list[4])

    # 使用したデータをエクセルファイルに記録する
    def data_excel_write(self):
        # 書き込むエクセルファイル
        book = openpyxl.Workbook()
        sheet = book.active

        # エクセルの一番最後の列に使用したデータのid, 日本語, 英語, 文化差ラベル, 類似度, hypernymを書き込む
        for line in self.test_data:
            max_row = sheet.max_row # シートの最後の行を取得
            sheet.cell(row = max_row + 1, column = 1).value = line.id
            sheet.cell(row = max_row + 1, column = 2).value = line.japanese
            sheet.cell(row = max_row + 1, column = 3).value = line.english
            sheet.cell(row = max_row + 1, column = 4).value = line.answer
            sheet.cell(row = max_row + 1, column = 5).value = line.similarity
            sheet.cell(row = max_row + 1, column = 6).value = line.hypernym
        # 保存
        book.save('../../研究/検証用データセット記録/1or1_dataset_hypernym10枚平均1118.xlsx')

    # 閾値ごとのAccuracyを計測するメソッド
    def threshold_by_accuracy(self, data_list, count):
        result = [] # 各閾値の結果を入れるリスト

        # 閾値ごとに計算
        for thre in self.thre_list:
            noculture_count = 0 # 文化差なしデータでの正解数をカウント
            culture_count = 0 # 文化差ありデータでの正解数をカウント

            # データのラベルと類似度をもとに閾値を使って正解数をカウント
            for data in data_list:
                # 文化差なしデータ
                if data.answer == '文化差なし':
                    # 類似度が閾値以上の場合、正解
                    if data.similarity >= thre:
                        noculture_count += 1
                # 文化差ありデータ
                elif '文化差あり' in data.answer:
                    # 類似度が閾値未満の場合、正解
                    if data.similarity < thre:
                        culture_count += 1

            # 閾値ごとに（文化差なしデータのAccuracy、文化差ありデータのAccuracy、全体でのAccuracy）を保存
            result.append([round(noculture_count/count, 2), round(culture_count/count, 2), round((noculture_count + culture_count)/count, 2)])

        return result

    # 結果をエクセルファイルに書き込むメソッド
    def writeresult(self):
        book = openpyxl.Workbook()
        sheet = book.active

        max_row = sheet.max_row # シートの最後の行を取得
        count = 1

        # ファイルの一番最後の行にを書き込む
        for thre in self.thre_list:
            sheet.cell(row = max_row + count, column = 1).value = thre
            sheet.cell(row = max_row + count, column=2).value = round(self.val_dict[thre][0],3)
            sheet.cell(row = max_row + count, column=3).value = round(self.val_dict[thre][1],3)
            sheet.cell(row = max_row + count, column=4).value = round(self.val_dict[thre][2],3)
            sheet.cell(row = max_row + count, column = 6).value = thre
            sheet.cell(row = max_row + count, column=7).value = round(self.test_dict[thre][0],3)
            sheet.cell(row = max_row + count, column=8).value = round(self.test_dict[thre][1],3)
            sheet.cell(row = max_row + count, column=9).value = round(self.test_dict[thre][2],3)
            count += 1
        # 保存
        book.save('../../研究/Accuracy結果/1or1_accuracy_hypernym10枚平均結果1118.xlsx')

    # 5分割交差検証のため、5回分の結果を平均する
    def average_result(self):
        self.val_dict = {key:[0,0,0] for key in self.thre_list}
        self.test_dict = {key:[0,0,0] for key in self.thre_list}
        
        for val_results in self.validation_result:
            for acc, thre in zip(val_results, self.thre_list):
                self.val_dict[thre][0] += acc[0]
                self.val_dict[thre][1] += acc[1]
                self.val_dict[thre][2] += acc[2]

        for test_results in self.test_result:
            for acc, thre in zip(test_results, self.thre_list):
                self.test_dict[thre][0] += acc[0]
                self.test_dict[thre][1] += acc[1]
                self.test_dict[thre][2] += acc[2]

        for thre in self.thre_list:
            self.val_dict[thre][0] /= 5
            self.test_dict[thre][0] /= 5
            self.val_dict[thre][1] /= 5
            self.test_dict[thre][1] /= 5
            self.val_dict[thre][2] /= 5
            self.test_dict[thre][2] /= 5

    # 結果を出力するメソッド
    def resultprint(self):
        count = 1
        for results in self.validation_result:
            print('{}回目'.format(count))
            for result, thre in zip(results, self.thre_list):
                print('閾値:{}, 文化差なし:{}, 文化差あり:{}, Total:{}'.format(thre, result[0], result[1], result[2]))
            count += 1
        count = 1
        for results in self.test_result:
            print('{}回目'.format(count))
            for result, thre in zip(results, self.thre_list):
                print('閾値:{}, 文化差なし:{}, 文化差あり:{}, Total:{}'.format(thre, result[0], result[1], result[2]))
            count += 1

# エクセルファイルをロード
#EXCEL_PATH = 'compare_vgg_kmeans.xlsx'
EXCEL_PATH = '../../研究/類似度結果/1or1多数決ラベル_hypernym10枚平均結果2022_1110.xlsx'
book = openpyxl.load_workbook(EXCEL_PATH)
sheet = book['Sheet1']

# インスタンスを生成
to = AccuracyCalculate()

# データを追加
for line in sheet.iter_rows(min_row=3):
    values = []
    for item in line:
        values.append(item.value)
    to.add(SuperExcelData(values[0], values[1], values[2], values[3], values[4], values[5]))

# 保存
book.save(EXCEL_PATH)
# 終了
book.close()

# 5分割交差検証のためにデータを下準備
to.dataset()
# 使用するデータをエクセルに記録
to.data_excel_write()

# 5グループ分のデータを取り出す
data_list = [to.data1, to.data2, to.data3, to.data4, to.data5]
# データ5つのうちから4つのデータをバリデーションに使って, バリデーションする 5C4の全ての組み合わせを出す
for datacomb in itertools.combinations(data_list, 4):
    validationdata = datacomb[0] + datacomb[1] + datacomb[2] + datacomb[3]
    to.validation_result.append(to.threshold_by_accuracy(validationdata, len(validationdata)))

# データの5つのうちから1つのデータをテストデータに使って, テストする
for data in reversed(data_list):
    to.test_result.append(to.threshold_by_accuracy(data, len(data)))

# 結果を出力
to.resultprint()
# ５回分の結果を平均
to.average_result()
# 結果をエクセルファイルに書き込む
to.writeresult()
