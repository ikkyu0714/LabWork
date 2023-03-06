"""
Excelに画像貼り付け.py
"""
import os
import glob
import imghdr
import openpyxl
import cv2

# 変数
max_height = [] # 各行の画像の高さの最大値を保持

class ExcelImageInsert():
    def __init__(self):
        self.readfile = 'synsetアンケートのコピー.xlsx' # 読み込むsynsetのリストを記録したファイル
        #self.filename='アンケート1.xlsx' # 結果を書き込むExcelのパス
        self.jpn_keyword_dict = {} # キーワードのリストを取得する辞書（日本語）
        self.eng_keyword_dict = {} # キーワードのリストを取得する辞書（英語）
        self.ind_keyword_dict = {}
        self.jpn_explanation_dict = {} #
        self.eng_explanation_dict = {}

    # Synsetを取得する
    def get_synset(self):
        key = 1
        book = openpyxl.load_workbook(self.readfile)
        sheet = book.worksheets[0]

        for row in sheet.iter_rows(min_row=2, min_col=2):
            value = []
            for cell in row:
                values = cell.value
                #print(values)
                values = values.replace(',', '')
                value.append(values.replace('_', ' '))

            self.jpn_keyword_dict[key] = value[0]
            self.eng_keyword_dict[key] = value[1]
            #self.ind_keyword_dict[key] = value[2]
            self.jpn_explanation_dict[key] = value[3]
            self.eng_explanation_dict[key] = value[4]
            key += 1

    def head_insert(self, type):
        if type == 'jpn':
            ws.cell(row = 1, column = 1).value = 'No.'
            ws.cell(row = 1, column = 2).value = '画像'
            ws.cell(row = 1, column = 3).value = '画像に写っているもの同士が同じものを指しているか（1:全く違う, 2:少しはあっている, 3:ある程度あっている, 4:おおかたあっている, 5:完璧にあっている）'
            ws.cell(row = 1, column = 5).value = '説明と画像がどのくらいあっているか（1:全く違う, 2:少しはあっている, 3:ある程度あっている, 4:おおかたあっている, 5:完璧にあっている）'
            ws.cell(row = 1, column = 7).value = 'キーワードと画像がどのくらい一致しているか（1:全く違う, 2:少しはあっている, 3:ある程度あっている, 4:おおかたあっている, 5:完璧にあっている）'
            ws.cell(row = 1, column = 9).value = 'キーワードと説明がどのくらい一致しているか（1:全く違う, 2:少しはあっている, 3:ある程度あっている, 4:おおかたあっている, 5:完璧にあっている）'
            ws.cell(row = 1, column = 11).value = 'キーワードと説明をみて、この概念を知っているかどうか（Yes or No）'
            ws.cell(row = 1, column = 12).value = '備考'
        elif type == 'eng':
            ws.cell(row = 1, column = 1).value = 'No.'
            ws.cell(row = 1, column = 2).value = 'Images'
            ws.cell(row = 1, column = 3).value = 'Rate the similarity between two sets of images（1: Dissimilar 2:Slightly dissimilar 3: Slightly similar 4: Similar 5: Same）'
            ws.cell(row = 1, column = 5).value = 'Rate the accuracy of the description for each image set（1: Dissimilar 2:Slightly dissimilar 3: Slightly similar 4: Similar 5: Same）'
            ws.cell(row = 1, column = 7).value = 'Rate the accuracy of the keyword(s) for each image set（1: Dissimilar 2:Slightly dissimilar 3: Slightly similar 4: Similar 5: Same）'
            ws.cell(row = 1, column = 9).value = 'Does the keywords match(es) the description?（1: Dissimilar 2:Slightly dissimilar 3: Slightly similar 4: Similar 5: Same）'
            ws.cell(row = 1, column = 11).value = 'Do you understand this concept by looking at the keywords and explanations?（Yes or No）'
            ws.cell(row = 1, column = 12).value = 'Remarks'
        elif type == 'indnesia':
            ws.cell(row = 1, column = 1).value = 'No.'
            ws.cell(row = 1, column = 2).value = 'Gambar'
            ws.cell(row = 1, column = 3).value = 'Nilai kemiripan antara dua kumpulan gambar（1: Berbeda 2: Sedikit berbeda 3: Sedikit mirip 4: Serupa 5: Sama）'
            ws.cell(row = 1, column = 5).value = 'Nilai keakuratan deskripsi untuk setiap kumpulan gambar（1: Berbeda 2: Sedikit berbeda 3: Sedikit mirip 4: Serupa 5: Sama）'
            ws.cell(row = 1, column = 7).value = 'Nilai keakuratan kata kunci untuk setiap kumpulan gambar（1: Berbeda 2: Sedikit berbeda 3: Sedikit mirip 4: Serupa 5: Sama）'
            ws.cell(row = 1, column = 9).value = 'Apakah kata kunci cocok atau sesuai dengan deskripsi?（1: Berbeda 2: Sedikit berbeda 3: Sedikit mirip 4: Serupa 5: Sama）'
            ws.cell(row = 1, column = 11).value = 'Apakah anda mengerti konsep ini dengan melihat kata kunci dan penjelasannya?（Yes or No）'
            ws.cell(row = 1, column = 12).value = 'Remarks'

    def get_file_names(self, set_dir_name):
        file_names = os.listdir(set_dir_name)
        temp_full_file_names = [os.path.join(set_dir_name, file_name) for file_name in file_names if os.path.isfile(os.path.join(set_dir_name, file_name))] # ファイルかどうかを判定
        return temp_full_file_names

    def attach_img(self, target_full_file_names, key, word, explanation, position, type):
        """
        画像を呼び出して、Excelに貼り付け
        """
        max_row = ws.max_row
        set_row_idx = max_row
        #print(set_row_idx)
        column_letter = 'B'
        #print(ws.cell(row=set_row_idx, column=1).column)
        #column_letter = ws.cell(row=set_row_idx, column=set_column_idx).column # セルの行列番号から、そのセルの列番号の文字列を取得
        #ws.cell(row=1, column=set_column_idx).value = set_dir_name # 各列の1行目に、貼り付ける画像があるディレクトリ名を入力
        merge_first = ws.cell(row=max_row + 2,column=9).coordinate
        merge_last = ws.cell(row=max_row + 3,column=9).coordinate
        #merge_last = ws.cell(row=max_row + 4,column=9).coordinate
        merge_range = merge_first + ':' + merge_last
        ws.merge_cells(merge_range)
        merge_first = ws.cell(row=max_row + 2,column=11).coordinate
        #merge_last = ws.cell(row=max_row + 4,column=11).coordinate
        merge_last = ws.cell(row=max_row + 3,column=11).coordinate
        merge_range = merge_first + ':' + merge_last
        ws.merge_cells(merge_range)
        merge_first = ws.cell(row=max_row + 2,column=3).coordinate
        merge_last = ws.cell(row=max_row + 3,column=3).coordinate
        merge_range = merge_first + ':' + merge_last
        ws.merge_cells(merge_range)
        ws.cell(row = max_row +2, column = 1).value = str(key) + '-(1)'
        ws.cell(row = max_row +3, column = 1).value = str(key) + '-(2)'
        #ws.cell(row = max_row+4, column = 1).value = str(key) + '-(3)'

        if type == 'jpn':
            #ws.cell(row = max_row+1, column = 3).value = '一番上画像(1)と画像(2), 真ん中画像(2)と(3), 一番下画像(1)と(3)'
            ws.cell(row = max_row+1, column = 5).value = '(説明) ' + explanation # 各列の1行目に、貼り付ける画像があるディレクトリ名を入力
            ws.cell(row = max_row+1, column = 7).value = '(キーワード) ' + word
        else:
            #ws.cell(row = max_row+1, column = 3).value = 'Top, image (1) and image (2), middle, image (2) and (3), bottom, image (1) and (3)'
            ws.cell(row = max_row+1, column = 5).value = '(Description) ' + explanation # 各列の1行目に、貼り付ける画像があるディレクトリ名を入力
            ws.cell(row = max_row+1, column = 7).value = '(Keyword) ' + word
        
        set_row = 2
        max_width = 0 # 画像の幅の最大値を保持するための変数
        #target_full_file_names.sort() # ファイル名でソート
        for target_file in target_full_file_names:
            if imghdr.what(target_file) != None: # 画像ファイルかどうかの判定
                img = openpyxl.drawing.image.Image(target_file)
                img.width = 400
                img.height = 200
                #print('[' + column_letter + '][' + str(set_row_idx+1) + ']' + target_file + 'を貼り付け')

                # 画像のサイズを取得して、セルの大きさを合わせる（画像同士が重ならないようにするため）
                size_img = cv2.imread(target_file)
                size_img_resize = cv2.resize(size_img, (430, 250))
                height, width = size_img_resize.shape[:2]
                if max_width < width:
                    max_width = width
                if not max_height[set_row_idx-1:set_row_idx]: # 配列「max_height」において、「set_row_idx」番目の要素が存在しなければ、挿入
                    max_height.insert(set_row_idx-1, height)
                if max_height[position-1] < height:
                    max_height[position-1] = height
                ws.row_dimensions[position+1].height = max_height[position-1] * 0.75
                ws.column_dimensions[column_letter].width = max_width * 0.13

                cell_address = ws.cell(row = max_row + set_row, column = 2).coordinate # セルの行列番号から、そのセルの番地を取得
                img.anchor = cell_address
                ws.add_image(img) # シートに画像貼り付け

                set_row += 1

            set_row_idx += 1

# 定数設定
INPUT_IMG_DIR = 'combine' # 貼り付ける画像を置いておくルートディレクトリ
SHEET_TITLE = '画像貼り付け' # シート名の設定
RESULT_FILE_NAME = 'synsetアンケートeng_4.xlsx' # 結果を保存するファイル名

eii = ExcelImageInsert()
eii.get_synset()
language = 'eng'

# ワークブック設定
wb = openpyxl.Workbook()
ws = wb.worksheets[0] # 1番目のシートを編集対象にする
ws.title = SHEET_TITLE # 1番目のシートに名前を設定

eii.head_insert(language)

# 貼り付ける画像を置いておくルートディレクトリ内のディレクトリ名を再帰的に取得
#dirs2 = glob.glob(os.path.join(os.path.join('combine', eii.jpn_keyword_dict[key]), '**' + os.sep), recursive=True)
pos = 1
for key in range(751,1001): # 1000個一気にやるとメモリ不足になるから100ずつ区切ってやる
    print(key)
    #print(eii.ind_keyword_dict[key])
    dirs = []
    jpn_root = os.path.join('combine', eii.jpn_keyword_dict[key])
    eng_root = os.path.join('combine', eii.eng_keyword_dict[key])
    #ind_root = os.path.join('combine', eii.ind_keyword_dict[key])
    dir_jpn = os.listdir(jpn_root)
    dir_eng = os.listdir(eng_root)
    #dir_ind = os.listdir(ind_root)
    if key % 2 == 1:
        dirs.append(os.path.join(jpn_root, dir_jpn[0]))
        dirs.append(os.path.join(eng_root, dir_eng[0]))
    elif key % 2 == 0:
        dirs.append(os.path.join(eng_root, dir_eng[0]))
        dirs.append(os.path.join(jpn_root, dir_jpn[0]))
    """if key % 3 == 1:
        dirs.append(os.path.join(jpn_root, dir_jpn[0]))
        dirs.append(os.path.join(eng_root, dir_eng[0]))
        dirs.append(os.path.join(ind_root, dir_ind[0]))
    elif key % 3 == 2:
        dirs.append(os.path.join(eng_root, dir_eng[0]))
        dirs.append(os.path.join(ind_root, dir_ind[0]))
        dirs.append(os.path.join(jpn_root, dir_jpn[0]))
    elif key % 3 == 0:
        dirs.append(os.path.join(ind_root, dir_ind[0]))
        dirs.append(os.path.join(jpn_root, dir_jpn[0]))
        dirs.append(os.path.join(eng_root, dir_eng[0]))"""

    if language == 'jpn':
        eii.attach_img(dirs, key, eii.jpn_keyword_dict[key], eii.jpn_explanation_dict[key], pos, language) # 画像貼り付け設定
    elif language == 'eng':
        eii.attach_img(dirs, key, eii.eng_keyword_dict[key], eii.eng_explanation_dict[key], pos, language) # 画像貼り付け設定
    #printeii.attach_img(dirs, key, eii.ind_keyword_dict[key], eii.eng_explanation_dict[key], pos) # 画像貼り付け設定


    pos += 1

# ファイルへの書き込み
wb.save(RESULT_FILE_NAME)
wb.close()