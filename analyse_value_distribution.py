import openpyxl

class CultureJudge():
    def __init__(self):
        self.list = []
        self.judge_list = []
        self.japanese_list = []
        self.indonesian_list = []
        self.choice_list = []
        self.searchword_list = ["object", "matter", "process", "causal_agent", "thing", "psychological_feature", "attribute", "group", "communication", "measure", "relation"]
        self.zerocount = 0
        self.count = 0
    
    def add(self, values):
        self.list.append(values)
        
    def print_data(self):
        for list in self.list:
            print(list)

    def distribution(self):
        japanese_dict = {"1":0, "2":0, "3":0, "4":0, "5":0}
        english_dict = {"1":0, "2":0, "3":0, "4":0, "5":0}
        for list in self.list:
            for value in list:
                if value == 1:
                    japanese_dict['1'] += 1
                elif value == 2:
                    japanese_dict['2'] += 1
                elif value == 3:
                    japanese_dict['3'] += 1
                elif value == 4:
                    japanese_dict['4'] += 1
                elif value == 5:
                    japanese_dict['5'] += 1
        print(japanese_dict)

    def choicedata(self, data):
        if 0 in data:
            self.zerocount += 1
        elif data[3] == 1 and data[5] == 1:
            self.count += 1
        elif data[8] == 1 and data[10] == 1:
            self.count += 1
        else:
            self.choice_list.append(data)

    def able_imagine(self):
        for line in self.choice_list:
            japanese = []
            japanese_dict = {"1":0, "2":0, "3":0, "4":0, "5":0}
            english_dict = {"1":0, "2":0, "3":0, "4":0, "5":0}
            value = 0
            japanese.append(line[0])
            if line[3] != 1 and line[5] != 1:
                value = (line[3]+line[5])//2
                japanese.append('想起できる({})'.format(value))
            elif line[3] == 1 or line[5] == 1:
                japanese.append('想起できない')
            else:
                japanese.append('想起できない')
            value = 0
            if line[7] != 1 and line[9] != 1:
                value = (line[7]+line[9])//2
                japanese.append('想起できる({})'.format(value))
            elif line[7] == 1 or line[9] == 1:
                japanese.append('想起できない')
            else:
                japanese.append('想起できない')
            self.japanese_list.append(japanese)
            value = 0
            indonesian = []
            indonesian.append(line[0])
            if line[4] != 1 and line[6] != 1:
                value = (line[4]+line[6])//2
                indonesian.append('想起できる({})'.format(value))
            elif line[4] == 1 or line[6] == 1:
                indonesian.append('想起できない')
            else:
                indonesian.append('想起できない')
            value = 0
            if line[8] != 1 and line[10] != 1:
                value = (line[8]+line[10])//2
                indonesian.append('想起できる({})'.format(value))
            elif line[8] == 1 or line[10] == 1:
                indonesian.append('想起できない')
            else:
                indonesian.append('想起できない')
            self.indonesian_list.append(indonesian)

    def judgeculture(self):
        for jpn, ind in zip(self.japanese_list, self.indonesian_list):
            list = []
            if jpn[0] == ind[0]:
                list.append(jpn[0])
                if jpn[1] == ind[1] and jpn[2] == ind[2]:
                    list.append('文化差なし')
                elif jpn[1] != ind[1] and jpn[2] == ind[2]:
                    list.append('文化差あり(排他関係)')
                else:
                    list.append('文化差あり(包含関係)')
            else:
                print('おかしいよ')
                exit()
            self.judge_list.append(list)

    def write_excel(self):
        book = openpyxl.Workbook()
        sheet = book.active
        row = 1
        number = 0
        while row <= 1000:
            data = self.judge_list[number]
            if data[0] == row:
                sheet.cell(row = row, column = 1).value = data[1]
                number += 1
                row += 1
            else:
                row += 1
        book.save('人手判定(日英)_HCI学会.xlsx')

    def printjudge(self):
        includecount = 0
        exclusivecount = 0
        nocount = 0
        for line in self.judge_list:
            if '包含' in line[1]:
                includecount += 1
            elif '排他' in line[1]:
                exclusivecount += 1
            elif line[1] == '文化差なし':
                nocount += 1
            print(line)
        print('排他関係:{}, 包含関係:{}, 文化差なし:{}, 文化差あり:{}'.format(exclusivecount, includecount, nocount, exclusivecount+includecount))

cj = CultureJudge()
books = 'まとめ結果thing.xlsx'
book = openpyxl.load_workbook(books)
sheet = book['Sheet']
for line in sheet.iter_rows(min_row=2):
    values = []
    for item in line:
        values.append(item.value)
    cj.add(values)
cj.print_data()
cj.distribution()