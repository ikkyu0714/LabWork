import openpyxl

class HumanImagine():
    def __init__(self):
        self.list = []
        self.imagine_list = []
        self.id_number = 1
        self.dict = {}
        self.count = 0

    # 想起できるか判別
    def judge(self, target_list):
        imagine_list = []
        for num in range(2):
            description_number = num + 1
            keyword_number = num + 3
            # target_list(画像同士, 説明(jpn), 説明(eng), キーワード(jpn), キーワード(eng), 説明とキーワード)
            if target_list[description_number] != 1 and target_list[keyword_number] == 1:
                if (target_list[description_number], target_list[keyword_number]) in self.dict:
                    self.dict[(target_list[description_number], target_list[keyword_number])] += 1
                else:
                    self.dict[(target_list[description_number], target_list[keyword_number])] = 1

            if target_list[description_number] == 0:
                imagine_list.append('知らない')
            # 説明とキーワードの一致度があるとき、説明の評価値のみを参照する
            elif target_list[5] != 1:
                if target_list[description_number] == 1 or target_list[keyword_number] == 1:
                    imagine_list.append('想起できない')
                else:
                    imagine_list.append('想起できる')
            # 説明とキーワードの評価値を参照して、想起できるか判断する
            elif target_list[description_number] == 1 and target_list[keyword_number] == 1:
                imagine_list.append('想起できない')
            elif target_list[description_number] != 1 or target_list[keyword_number] != 1:
                imagine_list.append('想起できる({})'.format(round((target_list[description_number] + target_list[keyword_number]) / 2)))
            else:
                imagine_list.append('想起できない')

        return imagine_list

    # 数値の値から想起できると想起できないを判定する
    def able_imagine(self):
        index_num_start = 0
        for index_num_start in range(0, len(self.list), 7):
            target_list = self.list[index_num_start:index_num_start+7]
            self.imagine_list.append(self.judge(target_list))
            self.id_number += 1

    def write_excel(self):
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.cell(row = 1, column = 1).value = '日本人_想起できるか'
        sheet.cell(row = 2, column = 1).value = '日本語画像'
        sheet.cell(row = 2, column = 2).value = '英語画像'
        for values in self.imagine_list:
            max_row = sheet.max_row # シートの最後の行を取得
            sheet.cell(row = max_row + 1, column = 1).value = values[0]
            sheet.cell(row = max_row + 1, column = 2).value = values[1]

        book.save('../../研究/アンケートデータ/想起データ/english想起データ多数決_1なし.xlsx')

hi = HumanImagine()
books = '../../研究/アンケートデータ/平均データ/アメリカ人平均.xlsx'
book = openpyxl.load_workbook(books)
sheet = book['Sheet']

for lines in sheet.iter_cols(min_row=2, min_col=1):
    for line in lines:
        hi.list.append(line.value)

hi.able_imagine()
hi.write_excel()