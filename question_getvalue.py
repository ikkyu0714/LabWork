import openpyxl
import os

class Questionaire_GetValue():
    def __init__(self):
        self.list = []
        self.row_count = 1

    def add(self, data):
        self.list.append(data)

    def row_add(self):
        self.row_count += 4

    def row_reset(self):
        self.row_count = 1

    def get_data(self):
        image_simi1 = sheet.cell(row = self.row_count + 2, column = 3).value
        image_simi2 = sheet.cell(row = self.row_count + 3, column = 3).value
        image_simi3 = sheet.cell(row = self.row_count + 4, column = 3).value
        des_simi1 = sheet.cell(row = self.row_count + 2, column = 5).value
        des_simi2 = sheet.cell(row = self.row_count + 3, column = 5).value
        des_simi3 = sheet.cell(row = self.row_count + 4, column = 5).value
        word_simi1 = sheet.cell(row = self.row_count + 2, column = 7).value
        word_simi2 = sheet.cell(row = self.row_count + 3, column = 7).value
        word_simi3 = sheet.cell(row = self.row_count + 4, column = 7).value
        correct_word = sheet.cell(row = self.row_count + 2, column = 9).value
        known_word = sheet.cell(row = self.row_count + 2, column = 11).value
        data = [image_simi1, image_simi2, image_simi3, des_simi1, des_simi2, des_simi3, word_simi1, word_simi2, word_simi3, correct_word, known_word]
        self.add(data)

    def print_list(self):
        for data in self.list:
            print(data)

    def write_excel(self):
        book = openpyxl.Workbook()
        sheet = book.active
        row = 1
        sheet.cell(row = row, column = 1).value = 'ID'
        sheet.cell(row = row, column = 2).value = '画像1vs2'
        sheet.cell(row = row, column = 3).value = '画像2vs3'
        sheet.cell(row = row, column = 4).value = '画像1vs3'
        sheet.cell(row = row, column = 5).value = '説明1'
        sheet.cell(row = row, column = 6).value = '説明2'
        sheet.cell(row = row, column = 7).value = '説明3'
        sheet.cell(row = row, column = 8).value = 'キーワード1'
        sheet.cell(row = row, column = 9).value = 'キーワード2'
        sheet.cell(row = row, column = 10).value = 'キーワード3'
        sheet.cell(row = row, column = 11).value = '説明とキーワード'
        sheet.cell(row = row, column = 12).value = '知っているか'
        for data in self.list:
            print(row)
            sheet.cell(row = row + 1, column = 1).value = row
            sheet.cell(row = row + 1, column = 2).value = data[0]
            sheet.cell(row = row + 1, column = 3).value = data[1]
            sheet.cell(row = row + 1, column = 4).value = data[2]
            sheet.cell(row = row + 1, column = 5).value = data[3]
            sheet.cell(row = row + 1, column = 6).value = data[4]
            sheet.cell(row = row + 1, column = 7).value = data[5]
            sheet.cell(row = row + 1, column = 8).value = data[6]
            sheet.cell(row = row + 1, column = 9).value = data[7]
            sheet.cell(row = row + 1, column = 10).value = data[8]
            sheet.cell(row = row + 1, column = 11).value = data[9]
            sheet.cell(row = row + 1, column = 12).value = data[10]
            row += 1
        book.save('../../研究/アンケートデータ/日本データ抜き出し/english1結果.xlsx')

language = 'eng'
qgv = Questionaire_GetValue()
if language == 'jpn':
    directory = '../../研究/アンケート結果/日本/服部アンケート'
    book_name = ['synsetアンケート1_jpn.xlsx', 'synsetアンケート2_jpn.xlsx', 'synsetアンケート3_jpn.xlsx', 'synsetアンケート4_jpn.xlsx']
elif language == 'eng':
    directory = '../../研究/アンケート結果/英語/英語１アンケート'
    book_name = ['synsetアンケート1_eng.xlsx', 'synsetアンケート2_eng.xlsx', 'synsetアンケート3_eng.xlsx', 'synsetアンケート4_eng.xlsx']
elif language == 'ind':
    directory = 'アンケート結果/インドネシア/suryadiアンケート'
    book_name = ['synsetアンケート1_ind.xlsx', 'synsetアンケート2_ind.xlsx', 'synsetアンケート3_ind.xlsx', 'synsetアンケート4_ind.xlsx']

for books in book_name:
    book = openpyxl.load_workbook(os.path.join(directory, books))
    sheet = book['画像貼り付け']

    for i in range(250):
        qgv.get_data()
        qgv.row_add()
    qgv.row_reset()

qgv.write_excel()

