"""
研究用 2022/10/17
アンケートの数字の組み合わせをカウントするプログラム
"""

import openpyxl
import os
import matplotlib.pyplot as plt

class AverageQuestion():
    def __init__(self):
        self.average_list = []
        self.list = []
        self.jpn_combinations_dict = {}
        self.jpn_desc_list = []

    def write_excel(self):
        book = openpyxl.Workbook()
        sheet = book.active
        row = 2
        sheet.cell(row = 1, column = 1).value = 'アメリカ人'
        for data in self.average_list:
            print(row)
            if data == 'スペース':
                row += 1
            else:
                sheet.cell(row = row, column = 1).value = data
                row += 1

        book.save('../../研究/アンケートデータ/平均データ/アメリカ人平均.xlsx')

    def data_average(self, list):
        person_num = len(list) // 7 # アンケート回答者の人数
        array_list = [[] for i in range(person_num)] # 各回答者のアンケートの回答を入れるリスト 5人の場合:[[１人目],[２人目],[３人目],[４人目],[５人目]]
        jpn_description_list = []

        # 回答者ごとに回答を分別する
        for num in range(len(list)):
            person = int(num % person_num) # 何人目の回答者の回答かを判別する値
            array_list[person].append(list[num])

        # Yes以外の回答(No)があった時、その回答は信用できないため、Noの回答者の値を0にする
        for target in range(len(array_list)):
            value = []
            if array_list[target][6] == 'yes' or array_list[target][6] == 'Yes':
                value.extend(array_list[target])
            else:
                value.extend([0,0,0,0,0,0,'No'])
            array_list[target] = value

        # 画像(jpn)と説明の結果の組み合わせの通り数をカウント
        for question_number in range(2, 6, 1): # ４問分ループを回す

            # 各回答者の回答を保持する配列
            values_list = []

            # 回答者の人数分ループを回す
            for value in array_list:

                # 0以外の値をリストに格納
                if value[question_number] != 0:
                    values_list.append(value[question_number])

            values_list = sorted(values_list)

            # 辞書内にない組み合わせならその組み合わせを追加、ある組み合わせなら値を操作する
            if str(values_list) in self.jpn_combinations_dict.keys():
                self.jpn_combinations_dict[str(values_list)] += 1
            else:
                self.jpn_combinations_dict[str(values_list)] = 1

    # 使うデータを選ぶ もし, Noが入っている場合、その概念におけるNoの回答者の回答は使用しない
    def choice_data(self, data):
        person_num = len(data) // 7
        array_list = [[] for i in range(person_num)]
        values = []
        for item in data:
            if type(item) == str:
                if ' ' in item:
                    item.replace(' ','')
                if '\n' in item:
                    item.replace('\n','')
            values.append(item)

        data = values
        self.data_average(data)
        #self.list.append(data)

aq = AverageQuestion()
books = '../../研究/アンケートデータ/まとめデータ/まとめ結果japanese一覧.xlsx'
book = openpyxl.load_workbook(books)
sheet = book['Sheet']
max_row = sheet.max_row # 最後の行を取得
number = 1 # 何番目かをカウント
row_start = 2 # 取得する最初の行
row_stop = 8 # 取得する最後の行
list = []
average_list = []

# 最初の行から最後の行を取得する １質問ずつ取得
while row_stop <= max_row:
    question = [] # 一つの概念の質問をまとめるリスト

    # startからstopまでの行をスライス
    for line in sheet.iter_rows(min_row=row_start, max_row=row_stop):
        values = [] # 1行の値をまとめるリスト

        # 行内の値を取り出してvaluesに入れる
        for item in line:
            values.append(item.value)
        question.extend(values)
    aq.choice_data(question)
    row_start += 7
    row_stop += 7
    number += 1
#aq.write_excel()

sort_list = sorted(aq.jpn_combinations_dict.items())
labels = [item[0] for item in sort_list]
values = [item[1] for item in sort_list]

print('組み合わせの種類:{}, データの数:{}'.format(len(labels), sum(values)))

plt.bar(labels, values)
plt.xticks(rotation=45)
plt.show()


book = openpyxl.Workbook()
sheet = book.active
sheet.cell(row = 1, column = 1).value = '組み合わせ'
sheet.cell(row = 1, column = 2).value = '個数'
for label, value in zip(labels, values):
    max_row = sheet.max_row # シートの最後の行を取得
    sheet.cell(row = max_row + 1, column = 1).value = label
    sheet.cell(row = max_row + 1, column = 2).value = value

book.save('../../研究/アンケートデータ/値の組み合わせ.xlsx')