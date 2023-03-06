"""
2021/04/13 研究用
類似度リスト.xlskからsynsetの日本語と英語のキーワードを拾ってくる
"""
import openpyxl
import os

# ファイルのパス
filepath = '../../類似度リスト.xlsx'
writepath = 'synset_list.xlsx'

num = 0　

# ファイルを開く
workbook = openpyxl.load_workbook(filepath) # 読み取るファイル(類似度リスト.xlsx)
writebook = openpyxl.load_workbook(writepath) #書き込むファイル(synset_list.xlsx)

# 作業シートを指定
sheet = workbook.worksheets[0]
writesheet = writebook.worksheets[0]

# 一番最後の列数を取得
max_row = sheet.max_row
target = writesheet.max_row

# 読み取りファイルから書込みファイルにデータを書き込む
while num+2 <= max_row:
    writesheet.cell(column=2, row = target + 1 + num).value = sheet.cell(column=2, row=3+num).value
    writesheet.cell(column=3, row = target + 1 + num).value = sheet.cell(column=3, row=3+num).value
    num += 1

# 保存
workbook.save(filepath)
writebook.save(writepath)
# 終了
workbook.close()
writebook.close()
